from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import pandas as pd
import uuid, re, io
from pathlib import Path
import openpyxl

app = FastAPI(title="Asistente Peticiones Almacenes")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Estado en memoria
catalog_df = None
catalog_loaded = False
catalog_path_default = Path("catalogue.xlsx")
plantilla_path_default = Path("plantilla_pedido.xlsx")  # provisional
catalogs = {}
requests_store = {}
matches = {}
cart = {}  # key: (ref,color,talla) -> qty

REF_PATTERN = re.compile(r"\[([^\]]+)\]")
PAREN_PATTERN = re.compile(r"\(([^)]*)\)")

def parse_ref(ref_str):
    if not isinstance(ref_str, str):
        return {"ref": None, "color": None, "talla": None}
    ref_match = REF_PATTERN.search(ref_str)
    paren_match = PAREN_PATTERN.search(ref_str)
    ref = ref_match.group(1).strip() if ref_match else None
    color = talla = None
    if paren_match:
        inside = paren_match.group(1).strip()
        if "," in inside:
            c, t = [p.strip() or None for p in inside.split(",", 1)]
            color, talla = c, t
        else:
            color = inside or None
    return {"ref": ref, "color": color, "talla": talla}

def read_table(uploaded: UploadFile) -> pd.DataFrame:
    name = uploaded.filename.lower()
    content = uploaded.file.read()
    if name.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(content))
    elif name.endswith((".xlsx", ".xls")):
        df = pd.read_excel(io.BytesIO(content))
    else:
        raise HTTPException(400, "Formato no soportado")
    if df.empty:
        raise HTTPException(400, "Archivo vacío")
    df.columns = df.columns.str.strip().str.title()
    return df

def load_catalog_file(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"No se encontró catalogue.xlsx en {path}")
    cat = pd.read_excel(path)
    cat.columns = cat.columns.str.strip().str.title()
    colmap = {c.lower(): c for c in cat.columns}
    ref_col = colmap.get("referencia") or list(cat.columns)[0]
    ean_col = colmap.get("ean") or colmap.get("codbarras")
    color_col = colmap.get("color")
    talla_col = colmap.get("talla")
    parsed = cat[ref_col].apply(parse_ref)
    cat["_ref"] = parsed.apply(lambda x: x["ref"])
    cat["_color"] = cat[color_col] if color_col else parsed.apply(lambda x: x["color"])
    cat["_talla"] = cat[talla_col] if talla_col else parsed.apply(lambda x: x["talla"])
    cat["_ean"] = cat[ean_col] if ean_col else None
    cat["_nombre"] = cat.get("Nombre", None)
    return cat

def ensure_catalog_loaded():
    global catalog_df, catalog_loaded
    if not catalog_loaded:
        catalog_df = load_catalog_file(catalog_path_default)
        catalog_loaded = True

def _export_df(df: pd.DataFrame, fmt: str, filename: str) -> StreamingResponse:
    if fmt == "csv":
        buff = io.StringIO()
        df.to_csv(buff, index=False)
        buff.seek(0)
        return StreamingResponse(
            iter([buff.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )
    elif fmt == "xlsx":
        buff = io.BytesIO()
        with pd.ExcelWriter(buff, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)
        buff.seek(0)
        return StreamingResponse(
            buff,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )
    else:
        raise HTTPException(400, "Formato no soportado (csv|xlsx)")

class MatchRequest(BaseModel):
    catalog_id: str
    request_id: str

class CartLine(BaseModel):
    ref: str
    color: str | None = None
    talla: str | None = None
    qty: int = 1

@app.on_event("startup")
def startup_event():
    ensure_catalog_loaded()

# --------- Importación ventas (independiente) -----------
@app.post("/catalog/upload")
async def upload_catalog(file: UploadFile = File(...)):
    df = read_table(file)
    cat_cols = {c.lower(): c for c in df.columns}
    ref_col = cat_cols.get("referencia") or list(df.columns)[0]
    ean_col = cat_cols.get("ean") or cat_cols.get("codbarras")
    color_col = cat_cols.get("color")
    talla_col = cat_cols.get("talla")
    parsed = df[ref_col].apply(parse_ref)
    df["_ref"] = parsed.apply(lambda x: x["ref"])
    df["_color"] = df[color_col] if color_col else parsed.apply(lambda x: x["color"])
    df["_talla"] = df[talla_col] if talla_col else parsed.apply(lambda x: x["talla"])
    df["_ean"] = df[ean_col] if ean_col else None
    cid = uuid.uuid4().hex[:12]
    catalogs[cid] = df
    return {"catalog_id": cid, "rows": len(df)}

@app.post("/request/upload")
async def upload_request(file: UploadFile = File(...)):
    df = read_table(file)
    prod_col = df.columns[0]
    if len(df.columns) >= 3:
        qty_col = df.columns[2]
    elif len(df.columns) >= 2:
        qty_col = df.columns[1]
    else:
        qty_col = None
    lower_cols = {c.lower(): c for c in df.columns}
    if "cantidad" in lower_cols:
        qty_col = lower_cols["cantidad"]

    parsed = df[prod_col].apply(parse_ref)
    df["_ref"] = parsed.apply(lambda x: x["ref"])
    df["_color"] = parsed.apply(lambda x: x["color"])
    df["_talla"] = parsed.apply(lambda x: x["talla"])
    df["_qty"] = df[qty_col] if qty_col else None

    rid = uuid.uuid4().hex[:12]
    requests_store[rid] = df
    return {"request_id": rid, "rows": len(df)}

@app.post("/match")
async def do_match(body: MatchRequest):
    cat_df = catalogs.get(body.catalog_id)
    req_df = requests_store.get(body.request_id)
    if cat_df is None or req_df is None:
        raise HTTPException(404, "Catalogo o petición no encontrados")
    merged = req_df.merge(
        cat_df,
        on=["_ref", "_color", "_talla"],
        how="left",
        suffixes=("_req", "_cat"),
    )
    merged["estado"] = merged["_ean"].notna().map({True: "encontrado", False: "no_encontrado"})
    not_found = merged[merged["_ean"].isna()].copy()
    mid = uuid.uuid4().hex[:12]
    matches[mid] = {"merged": merged, "not_found": not_found}
    return {
        "match_id": mid,
        "total": len(merged),
        "encontrados": int(merged["_ean"].notna().sum()),
        "no_encontrados": int(not_found.shape[0]),
        "preview": merged.head(20).to_dict(orient="records"),
    }

@app.get("/match/{match_id}/export")
async def export_match(match_id: str, format: str = "xlsx", type: str = "all"):
    match = matches.get(match_id)
    if match is None:
        raise HTTPException(404, "Match no encontrado")
    df = match["not_found"] if type == "missing" else match["merged"]
    suffix = "missing" if type == "missing" else "all"
    filename = f"match_{match_id}_{suffix}"
    return _export_df(df, format.lower(), filename)

# --------- Búsqueda catálogo -----------
@app.get("/products/search")
async def search_products(q: str):
    ensure_catalog_loaded()
    q_lower = q.lower()
    mask = catalog_df["_ref"].str.lower().str.contains(q_lower, na=False)
    if "_nombre" in catalog_df.columns and catalog_df["_nombre"] is not None:
        mask = mask | catalog_df["_nombre"].fillna("").str.lower().str.contains(q_lower, na=False)
    found = catalog_df[mask].copy()
    results = {}
    for _, row in found.iterrows():
        ref = row["_ref"]
        nombre = row.get("_nombre", None)
        key = (ref, nombre)
        if key not in results:
            results[key] = {"ref": ref, "nombre": nombre, "variantes": []}
        results[key]["variantes"].append({
            "color": row["_color"],
            "talla": row["_talla"],
            "ean": row["_ean"],
        })
    return list(results.values())

# --------- Carrito manual -----------
@app.post("/cart/add")
async def cart_add(line: CartLine):
    ensure_catalog_loaded()
    key = (line.ref, line.color, line.talla)
    cart[key] = cart.get(key, 0) + line.qty
    if cart[key] <= 0:
        cart.pop(key, None)
    return {"ok": True, "items": len(cart)}

@app.post("/cart/remove")
async def cart_remove(line: CartLine):
    key = (line.ref, line.color, line.talla)
    if key in cart:
        cart[key] -= line.qty
        if cart[key] <= 0:
            cart.pop(key, None)
    return {"ok": True, "items": len(cart)}

@app.get("/cart/view")
async def cart_view():
    ensure_catalog_loaded()
    rows = []
    for (ref, color, talla), qty in cart.items():
        row = catalog_df[
            (catalog_df["_ref"] == ref)
            & (catalog_df["_color"] == color)
            & (catalog_df["_talla"] == talla)
        ]
        ean = row["_ean"].iloc[0] if len(row) else None
        nombre = row["_nombre"].iloc[0] if len(row) and "_nombre" in row.columns else None
        rows.append({
            "ref": ref,
            "color": color,
            "talla": talla,
            "ean": ean,
            "qty": qty,
            "nombre": nombre,
        })
    return {"items": rows}

# --------- Checkout con metadatos y plantilla -----------
@app.get("/cart/checkout")
async def cart_checkout(
    format: str = "xlsx",
    origin: str = "",
    destination: str = "",
    fecha: str = "",
    pedido_ref: str = ""
):
    ensure_catalog_loaded()
    data = []
    for (ref, color, talla), qty in cart.items():
        row = catalog_df[
            (catalog_df["_ref"] == ref)
            & (catalog_df["_color"] == color)
            & (catalog_df["_talla"] == talla)
        ]
        ean = row["_ean"].iloc[0] if len(row) else None
        data.append({
            "Origen": origin,
            "Destino": destination,
            "Fecha": fecha,
            "Pedido": pedido_ref,
            "Ref": ref,
            "Color": color,
            "Talla": talla,
            "EAN": ean,
            "Cantidad": qty,
            "estado": "encontrado" if ean else "no_encontrado",
        })
    merged = pd.DataFrame(data)
    not_found = merged[merged["EAN"].isna()] if not merged.empty else pd.DataFrame()

    if format.lower() == "csv":
        return _export_df(merged, "csv", "cart_checkout")

    # XLSX con plantilla
    if not plantilla_path_default.exists():
        return _export_df(merged, "xlsx", "cart_checkout")

    wb = openpyxl.load_workbook(plantilla_path_default)
    ws = wb.active  # única hoja (primera)

    start_row = 2  # datos desde fila 2
    row_idx = start_row
    for _, r in merged.iterrows():
        ws.cell(row=row_idx, column=1, value=r.get("Fecha"))          # A: Fecha
        ws.cell(row=row_idx, column=2, value=r.get("Origen"))         # B: Almacén de origen
        ws.cell(row=row_idx, column=3, value=r.get("Destino"))        # C: Almacén de destino
        ws.cell(row=row_idx, column=4, value=pedido_ref)              # D: Observaciones = referencia de pedido
        ws.cell(row=row_idx, column=5, value=r.get("EAN"))            # E: EAN
        ws.cell(row=row_idx, column=6, value=r.get("Cantidad"))       # F: Cantidad
        row_idx += 1

    buff = io.BytesIO()
    wb.save(buff)
    buff.seek(0)
    return StreamingResponse(
        buff,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="cart_checkout.xlsx"'},
    )

# Mount static files to serve UI from root path
# This is placed after all API routes to ensure API routes take precedence
# Serves index.html and other assets from static/ directory on port 8000


from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
# ... deja intactos los endpoints anteriores ...

@app.get("/", include_in_schema=False)
async def serve_index():
    return FileResponse(Path(__file__).parent / "static" / "index.html")

app.mount("/", StaticFiles(directory="static", html=True), name="static")
